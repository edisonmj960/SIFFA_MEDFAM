import os
import json
import csv
import io
import secrets
from datetime import date, datetime
from urllib.parse import urlencode
from flask import Flask, request, Response, redirect
from siifa_bulk_client import SiifaClient, SiifaApiError

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None


app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", secrets.token_hex(32))

_SESSIONS: dict[str, dict] = {}


_LOGO_MEDFAM_PATH = os.path.join(os.path.dirname(__file__), "logo_medfam.jpg")


@app.route("/assets/logo_medfam.jpg", methods=["GET"])
def assets_logo_medfam():
    try:
        with open(_LOGO_MEDFAM_PATH, "rb") as f:
            data = f.read()
    except Exception:
        return Response("Logo no disponible", status=404, mimetype="text/plain; charset=utf-8")
    return Response(data, mimetype="image/jpeg")


def _get_base_urls():
    return (
        os.environ.get("SIIFA_SECURITY_BASEURL", "https://siifa.sispro.gov.co/siifa-seguridad"),
        os.environ.get("SIIFA_FACTURA_BASEURL", "https://siifa.sispro.gov.co/siifa-factura"),
    )


def _get_session_id() -> str | None:
    sid = request.cookies.get("siifa_session")
    if not sid:
        return None
    if sid not in _SESSIONS:
        return None
    return sid


def _require_token() -> str | None:
    sid = _get_session_id()
    if not sid:
        return None
    token = _SESSIONS.get(sid, {}).get("token")
    if not token:
        return None
    return token


def _make_client_with_token(token: str) -> SiifaClient:
    seguridad_base_url, factura_base_url = _get_base_urls()
    client = SiifaClient(
        seguridad_base_url=seguridad_base_url,
        factura_base_url=factura_base_url,
    )
    client.token = token
    return client


def _get_session() -> dict | None:
    sid = _get_session_id()
    if not sid:
        return None
    return _SESSIONS.get(sid)


def _set_session_value(key: str, value):
    sess = _get_session()
    if sess is None:
        return
    sess[key] = value


def _get_session_value(key: str, default=None):
    sess = _get_session()
    if sess is None:
        return default
    return sess.get(key, default)


def _read_uploaded_text(file_storage) -> str:
    raw = file_storage.read()
    if raw is None:
        return ""
    if isinstance(raw, str):
        return raw
    return raw.decode("utf-8-sig", errors="replace")


def _parse_uploaded_json(file_storage):
    text = _read_uploaded_text(file_storage).strip()
    if not text:
        raise ValueError("Archivo vacío")
    return json.loads(text)


def _parse_uploaded_csv(file_storage) -> list[dict]:
    text = _read_uploaded_text(file_storage)
    if not text.strip():
        raise ValueError("Archivo vacío")
    sample = text[:4096]
    dialect = None
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        dialect = csv.get_dialect("excel")
    fp = io.StringIO(text)
    reader = csv.DictReader(fp, dialect=dialect)
    rows = []
    for row in reader:
        if not row:
            continue
        rows.append({(k or "").strip(): (v or "").strip() for k, v in row.items()})
    return rows


def _parse_uploaded_xlsx(file_storage) -> list[dict]:
    if load_workbook is None:
        raise ValueError("No está disponible la lectura de Excel (.xlsx) en este entorno.")
    raw = file_storage.read()
    if not raw:
        raise ValueError("Archivo vacío")
    wb = load_workbook(filename=io.BytesIO(raw), data_only=True)
    ws = wb.active

    headers = []
    for c in ws[1]:
        v = c.value
        headers.append(str(v).strip() if v is not None else "")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        out = {}
        for idx, h in enumerate(headers):
            if not h:
                continue
            out[h] = row[idx] if idx < len(row) else None
        rows.append(out)
    return rows


def _normalize_key(key: str) -> str:
    return "".join(ch for ch in (key or "").strip().lower() if ch.isalnum())


def _row_get(row: dict, *keys: str):
    norm = {_normalize_key(k): v for k, v in (row or {}).items()}
    for k in keys:
        nk = _normalize_key(k)
        if nk in norm:
            return norm[nk]
    return None


def _coerce_int(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return int(s)
    except Exception:
        return None


def _coerce_float(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _to_iso_z(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%dT%H:%M:%SZ")
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).strftime("%Y-%m-%dT00:00:00Z")
    s = str(value).strip()
    if not s:
        return None
    if "T" in s:
        return s
    return f"{s}T00:00:00Z"


def _xlsx_bytes(sheet_name: str, headers: list[str], rows: list[dict]) -> bytes:
    if Workbook is None:
        raise ValueError("No está disponible la exportación a Excel (.xlsx) en este entorno.")
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Hoja1")[:31]
    ws.append(list(headers))
    for r in rows:
        ws.append([r.get(h) for h in headers])
    fp = io.BytesIO()
    wb.save(fp)
    return fp.getvalue()


def _xlsx_bytes_multi(sheets: list[tuple[str, list[str], list[dict]]]) -> bytes:
    if Workbook is None:
        raise ValueError("No está disponible la exportación a Excel (.xlsx) en este entorno.")
    wb = Workbook()
    first = True
    for sheet_name, headers, rows in sheets:
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()
        ws.title = (sheet_name or "Hoja")[:31]
        ws.append(list(headers))
        for r in rows:
            ws.append([r.get(h) for h in headers])
    fp = io.BytesIO()
    wb.save(fp)
    return fp.getvalue()


def _factura_by_id(client: SiifaClient, id_factura: int) -> dict:
    try:
        fac = client.get_factura(int(id_factura))
        return fac if isinstance(fac, dict) else {}
    except Exception:
        pass
    try:
        page = client.list_facturas(IdFactura=int(id_factura), NumeroPagina=1, RegistrosPorPagina=1)
        if isinstance(page, dict):
            items = page.get("resultado") or []
            if items and isinstance(items[0], dict):
                return items[0]
    except Exception:
        pass
    return {}


def _facturas_map_by_ids(client: SiifaClient, ids: set[int], max_pages: int = 10) -> dict[int, dict]:
    ids = {int(x) for x in (ids or set()) if x is not None}
    if not ids:
        return {}
    found: dict[int, dict] = {}
    page_num = 1
    while page_num <= max_pages and len(found) < len(ids):
        try:
            page = client.list_facturas(NumeroPagina=page_num, RegistrosPorPagina=1500)
        except Exception:
            break
        if not isinstance(page, dict):
            break
        items = page.get("resultado") or []
        if not items:
            break
        for it in items:
            if not isinstance(it, dict):
                continue
            fid = it.get("idFactura")
            try:
                fid = int(fid)
            except Exception:
                continue
            if fid not in ids or fid in found:
                continue
            emisor = it.get("emisor") if isinstance(it.get("emisor"), dict) else {}
            found[fid] = {
                "numeroFactura": it.get("numeroFactura"),
                "nitEmisor": (emisor or {}).get("nitEmisor"),
            }
        total_pages = page.get("totalPaginas")
        if total_pages is not None:
            try:
                if page_num >= int(total_pages):
                    break
            except Exception:
                pass
        page_num += 1
    return found


def _resolve_id_factura(client: SiifaClient, row: dict) -> tuple[int | None, dict]:
    id_factura = _coerce_int(_row_get(row, "idFactura", "IdFactura"))
    if id_factura:
        return id_factura, {}

    numero = (str(_row_get(row, "numeroFactura", "NumeroFactura", "NUMERO DE FACTURA") or "").strip() or None)
    nit_emisor = (str(_row_get(row, "nitEmisor", "NitEmisor", "NUMERO IDENT PRESTADOR") or "").strip() or None)
    nit_adq = (str(_row_get(row, "nitAdquiriente", "NitAdquiriente", "NIT ADQUIRIENTE") or "").strip() or None)

    if not numero:
        return None, {"warning": "Sin idFactura ni numeroFactura"}

    page = client.list_facturas(
        NumeroFactura=numero,
        NitEmisor=nit_emisor,
        NitAdquiriente=nit_adq,
        RegistrosPorPagina=5,
        NumeroPagina=1,
    )
    items = page.get("resultado") or []
    if not items:
        return None, {"warning": "Factura no encontrada"}
    it = items[0] or {}
    return _coerce_int(it.get("idFactura")), {"warning": "Resuelta por NumeroFactura (primer match)" if len(items) > 1 else None}


LOGIN_HTML = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Ingreso - MedFam SIIFA</title>
  <style>{{ base_css|safe }}</style>
</head>
<body>
  <main class="container">
    <div class="page-title">
      <h1>Ingreso</h1>
      <div class="meta">Autenticación para consultar la API.</div>
    </div>
    <div class="card" style="max-width: 540px;">
      <div style="display:flex; gap:12px; align-items:center; margin-bottom: 10px;">
        <img class="brand-logo" src="/assets/logo_medfam.jpg" alt="MedFam" />
        <div>
          <div style="font-weight: 900; font-size: 18px;">MedFam</div>
          <div class="meta">Sistema SIIFA (consultas, cargues y respuestas).</div>
        </div>
      </div>
      {% if error %}
        <div class="error">
          <strong>Error:</strong> {{ error }}
          {% if details %}<pre>{{ details }}</pre>{% endif %}
        </div>
      {% endif %}
      <form method="post">
        <label>
          Usuario
          <input name="userName" autocomplete="username" required />
        </label>
        <label>
          Contraseña
          <input name="password" type="password" autocomplete="current-password" required />
        </label>
        <div class="actions">
          <button type="submit">Ingresar</button>
        </div>
      </form>
    </div>
    {{ footer|safe }}
  </main>
</body>
</html>
"""


BASE_CSS = """
:root{
  --bg:#f6f9ff;
  --surface:#ffffff;
  --card:#ffffff;
  --border:#dbe5f2;
  --text:#0f172a;
  --muted:#475569;
  --primary:#0ea5e9;
  --primary-2:#2563eb;
  --accent:#84cc16;
  --danger:#dc2626;
  --shadow:0 10px 25px rgba(15,23,42,.08);
}
*{ box-sizing:border-box; }
body{ margin:0; font-family:system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif; background:linear-gradient(180deg, var(--bg), #ffffff); color:var(--text); }
a{ color:var(--primary-2); text-decoration:none; }
.topbar{ position:sticky; top:0; z-index:50; background:rgba(255,255,255,.92); border-bottom:1px solid var(--border); }
.topbar-inner{ max-width:1200px; margin:0 auto; padding:12px 16px; display:flex; gap:12px; align-items:center; }
.brand{ display:flex; gap:10px; align-items:center; font-weight:900; letter-spacing:.2px; color:var(--text); }
.brand-logo{ height:34px; width:auto; border-radius:10px; box-shadow:0 6px 14px rgba(2,132,199,.12); background:#fff; border:1px solid rgba(148,163,184,.25); padding:4px; }
.brand-text{ color:var(--text); }
.navlinks{ display:flex; gap:10px; flex-wrap:wrap; align-items:center; }
.navlinks a{ color:var(--muted); padding:8px 10px; border-radius:12px; border:1px solid transparent; background:transparent; }
.navlinks a.active{ color:var(--primary-2); background:rgba(14,165,233,.12); border-color:rgba(14,165,233,.25); }
.spacer{ margin-left:auto; }
.logout{ color:var(--danger) !important; border-color:rgba(220,38,38,.25) !important; background:rgba(220,38,38,.06); }
.container{ max-width:1200px; margin:0 auto; padding:18px 16px 22px; }
.page-title{ display:flex; flex-wrap:wrap; gap:10px; align-items:flex-end; justify-content:space-between; margin:6px 0 14px; }
.page-title h1{ margin:0; font-size:28px; }
.meta{ color:var(--muted); font-size:13px; }
.card{ background:var(--card); border:1px solid var(--border); border-radius:16px; padding:14px; box-shadow:var(--shadow); }
.error{ background:rgba(220,38,38,.08); border:1px solid rgba(220,38,38,.25); border-radius:14px; padding:12px; margin-bottom:12px; }
form{ display:grid; grid-template-columns:repeat(auto-fit, minmax(220px, 1fr)); gap:12px; }
label{ display:grid; gap:6px; font-size:13px; color:var(--muted); }
input, select, button{ padding:10px 12px; font-size:14px; border-radius:12px; border:1px solid var(--border); background:rgba(255,255,255,.92); color:var(--text); }
input::placeholder{ color:rgba(71,85,105,.65); }
button{ background:linear-gradient(180deg, rgba(14,165,233,.18), rgba(37,99,235,.14)); border-color:rgba(14,165,233,.35); cursor:pointer; font-weight:750; }
button:hover{ background:linear-gradient(180deg, rgba(14,165,233,.24), rgba(37,99,235,.18)); }
.actions{ grid-column:1 / -1; display:flex; flex-wrap:wrap; gap:10px; align-items:center; }
.grid{ overflow:auto; max-height:62vh; border-radius:14px; border:1px solid var(--border); background:var(--surface); }
table{ width:100%; border-collapse:collapse; background:var(--surface); }
th, td{ border-bottom:1px solid rgba(15,23,42,.08); padding:10px; font-size:14px; vertical-align:top; }
th{ position:sticky; top:0; background:rgba(248,250,252,.95); text-align:left; color:rgba(15,23,42,.72); }
pre{ overflow:auto; background:rgba(248,250,252,.95); padding:10px; border:1px solid var(--border); border-radius:12px; }
.pager{ display:flex; gap:10px; align-items:center; margin:12px 0; flex-wrap:wrap; }
.footer{ margin-top:18px; padding:12px 0 4px; color:rgba(71,85,105,.8); font-size:13px; border-top:1px solid rgba(15,23,42,.08); }
"""


NAV_HTML = """
<header class="topbar">
  <div class="topbar-inner">
    <a class="brand" href="/">
      <img class="brand-logo" src="/assets/logo_medfam.jpg" alt="MedFam" />
      <span class="brand-text">MedFam</span>
    </a>
    <nav class="navlinks">
      <a href="/" class="{{ 'active' if active == 'facturas' else '' }}">Facturas</a>
      <a href="/seguimiento" class="{{ 'active' if active == 'seguimiento' else '' }}">Seguimiento</a>
      <a href="/responder-glosas" class="{{ 'active' if active == 'responder' else '' }}">Responder glosas</a>
      <a href="/pagos" class="{{ 'active' if active == 'pagos' else '' }}">Pagos</a>
      <a href="/consulta-masiva" class="{{ 'active' if active == 'consulta' else '' }}">Consultas masivas</a>
      <a href="/cargue" class="{{ 'active' if active == 'cargue' else '' }}">Cargue masivo</a>
    </nav>
    <div class="spacer"></div>
    <a class="navlinks logout" href="/logout">Cerrar sesión</a>
  </div>
</header>
"""


FOOTER_HTML = """
<footer class="footer">
  Desarrollado por Ing Edison Mejia - 2026
</footer>
"""


def _render_footer() -> str:
    return FOOTER_HTML


def _render_nav(active: str) -> str:
    from jinja2 import Template
    return Template(NAV_HTML).render(active=active)


HTML = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Consultas SIIFA - IPS Medfam</title>
  <style>
    {{ base_css|safe }}
  </style>
</head>
<body>
  {{ nav|safe }}
  <main class="container">
  <div class="page-title">
    <h1>Facturas</h1>
    <div class="meta">Consulta y exportación (JSON/CSV/Excel).</div>
  </div>
  {% if error %}
    <div class="error">
      <strong>Error:</strong> {{ error }}
      {% if details %}<pre>{{ details }}</pre>{% endif %}
    </div>
  {% endif %}
  <form method="get">
    <label>
      IdFactura
      <input name="IdFactura" value="{{ q.IdFactura or '' }}" />
    </label>
    <label>
      NumeroFactura
      <input name="NumeroFactura" value="{{ q.NumeroFactura or '' }}" />
    </label>
    <label>
      NitEmisor
      <input name="NitEmisor" value="{{ q.NitEmisor or '900243869' }}" />
    </label>
    <label>
      NitAdquiriente
      <input name="NitAdquiriente" value="{{ q.NitAdquiriente or '' }}" />
    </label>
    <label>
      FechaEmisionInicio (YYYY-MM-DDTHH:MM:SSZ)
      <input name="FechaEmisionInicio" value="{{ q.FechaEmisionInicio or '' }}" />
    </label>
    <label>
      FechaEmisionFinal (YYYY-MM-DDTHH:MM:SSZ)
      <input name="FechaEmisionFinal" value="{{ q.FechaEmisionFinal or '' }}" />
    </label>
    <label>
      FechaCargue (YYYY-MM-DDTHH:MM:SSZ)
      <input name="FechaCargue" value="{{ q.FechaCargue or '' }}" />
    </label>
    <label>
      TieneRadicado
      <select name="TieneRadicado">
        <option value="" {% if not q.TieneRadicado %}selected{% endif %}>Cualquiera</option>
        <option value="true" {% if q.TieneRadicado == 'true' %}selected{% endif %}>Con radicado</option>
        <option value="false" {% if q.TieneRadicado == 'false' %}selected{% endif %}>Sin radicado</option>
      </select>
    </label>
    <label>
      RegistrosPorPagina (1-1500)
      <input name="RegistrosPorPagina" type="number" min="1" max="1500" value="{{ q.RegistrosPorPagina or 50 }}" />
    </label>
    <label>
      NumeroPagina
      <input name="NumeroPagina" type="number" min="1" value="{{ q.NumeroPagina or 1 }}" />
    </label>
    <div style="grid-column: 1 / -1;">
      <button type="submit">Buscar</button>
      <button type="submit" name="descargar" value="json">Descargar JSON</button>
      <button type="submit" name="descargar" value="csv">Descargar CSV</button>
      <button type="submit" name="descargar" value="xlsx">Descargar Excel</button>
    </div>
  </form>

  {% if page %}
    <div class="meta">
      Total registros: {{ page.totalRegistros or 'N/D' }} |
      Total páginas: {{ page.totalPaginas or 'N/D' }} |
      Página actual: {{ page.paginaActual or q.NumeroPagina or 1 }}
    </div>
    <div class="pager">
      {% set prev = (q.NumeroPagina|int - 1) if q.NumeroPagina else 1 %}
      {% set next = (q.NumeroPagina|int + 1) if q.NumeroPagina else 2 %}
      <a href="/?{{ prev_q }}">« Anterior</a>
      <a href="/?{{ next_q }}">Siguiente »</a>
    </div>
    <div class="grid">
      <table>
        <thead>
          <tr>
            <th>IdFactura</th>
            <th>NumeroFactura</th>
            <th>CUFE</th>
            <th>TipoFactura</th>
            <th>NitEmisor</th>
            <th>Emisor</th>
            <th>NitAdquiriente</th>
            <th>Adquiriente</th>
            <th>FechaEmision</th>
            <th>TotalValorBruto</th>
            <th>ValorFactura</th>
          </tr>
        </thead>
        <tbody>
          {% for it in page.resultado or [] %}
            <tr>
              <td>{{ it.idFactura }}</td>
              <td>{{ it.numeroFactura }}</td>
              <td style="max-width: 420px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;">{{ it.cufe }}</td>
              <td>{{ it.tipoFactura }}</td>
              <td>{{ (it.emisor or {}).get('nitEmisor') }}</td>
              <td>{{ (it.emisor or {}).get('razonSocial') }}</td>
              <td>{{ (it.adquiriente or {}).get('nitAdquiriente') }}</td>
              <td>{{ (it.adquiriente or {}).get('razonSocial') }}</td>
              <td>{{ it.fechaEmision }}</td>
              <td>{{ it.totalValorBruto }}</td>
              <td>{{ it.valorFactura }}</td>
            </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
    <div class="pager">
      <a href="/?{{ prev_q }}">« Anterior</a>
      <a href="/?{{ next_q }}">Siguiente »</a>
    </div>
  {% endif %}
  {{ footer|safe }}
  </main>
</body>
</html>
"""


def _clean_query(args):
    q = {
        "IdFactura": args.get("IdFactura") or None,
        "NumeroFactura": args.get("NumeroFactura") or None,
        "NitEmisor": args.get("NitEmisor") or "900243869",
        "NitAdquiriente": args.get("NitAdquiriente") or None,
        "FechaEmisionInicio": args.get("FechaEmisionInicio") or None,
        "FechaEmisionFinal": args.get("FechaEmisionFinal") or None,
        "FechaCargue": args.get("FechaCargue") or None,
        "TieneRadicado": args.get("TieneRadicado") or None,
        "RegistrosPorPagina": min(1500, max(1, int(args.get("RegistrosPorPagina") or 50))),
        "NumeroPagina": int(args.get("NumeroPagina") or 1),
    }
    if q["IdFactura"] is not None:
        try:
            q["IdFactura"] = int(str(q["IdFactura"]).strip())
        except ValueError:
            q["IdFactura"] = None
    if q["TieneRadicado"] not in (None, "", "true", "false"):
        q["TieneRadicado"] = None
    return q


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    details = None
    if request.method == "POST":
        user = (request.form.get("userName") or "").strip()
        pwd = request.form.get("password") or ""
        try:
            seguridad_base_url, factura_base_url = _get_base_urls()
            client = SiifaClient(seguridad_base_url=seguridad_base_url, factura_base_url=factura_base_url)
            token = client.login(user, pwd)
            sid = secrets.token_urlsafe(32)
            _SESSIONS[sid] = {"token": token, "userName": user}
            resp = redirect("/")
            resp.set_cookie("siifa_session", sid, httponly=True, samesite="Lax")
            return resp
        except SiifaApiError as e:
            if e.status == 401:
                sid = request.cookies.get("siifa_session")
                if sid:
                    _SESSIONS.pop(sid, None)
                resp = redirect("/login")
                resp.delete_cookie("siifa_session")
                return resp
            error = str(e)
            details = json.dumps(e.payload, ensure_ascii=False, indent=2) if e.payload is not None else None
        except Exception as e:
            error = str(e)

    from jinja2 import Template
    tmpl = Template(LOGIN_HTML)
    body = tmpl.render(error=error, details=details, base_css=BASE_CSS, footer=_render_footer())
    return Response(body, mimetype="text/html; charset=utf-8")


@app.route("/logout", methods=["GET"])
def logout():
    sid = request.cookies.get("siifa_session")
    if sid and sid in _SESSIONS:
        _SESSIONS.pop(sid, None)
    resp = redirect("/login")
    resp.delete_cookie("siifa_session")
    return resp


@app.route("/", methods=["GET"])
def index():
    token = _require_token()
    if not token:
        return redirect("/login")

    args = request.args
    q = _clean_query(args)
    page = None
    error = None
    details = None

    if any(v for k, v in q.items() if k not in ("RegistrosPorPagina", "NumeroPagina")) or args:
        try:
            client = _make_client_with_token(token)
            tiene = q["TieneRadicado"]
            if tiene == "true":
                tiene_bool = True
            elif tiene == "false":
                tiene_bool = False
            else:
                tiene_bool = None

            page = client.list_facturas(
                IdFactura=q["IdFactura"],
                NumeroFactura=q["NumeroFactura"],
                NitEmisor=q["NitEmisor"],
                NitAdquiriente=q["NitAdquiriente"],
                FechaEmisionInicio=q["FechaEmisionInicio"],
                FechaEmisionFinal=q["FechaEmisionFinal"],
                TieneRadicado=tiene_bool,
                FechaCargue=q["FechaCargue"],
                NumeroPagina=q["NumeroPagina"],
                RegistrosPorPagina=q["RegistrosPorPagina"],
            )

            descargar = args.get("descargar")
            if descargar == "json":
                return Response(json.dumps(page, ensure_ascii=False, indent=2), mimetype="application/json")
            if descargar == "csv":
                fp = io.StringIO()
                writer = csv.writer(fp)
                writer.writerow(
                    [
                        "idFactura",
                        "numeroFactura",
                        "cufe",
                        "tipoFactura",
                        "nitEmisor",
                        "razonSocialEmisor",
                        "nitAdquiriente",
                        "razonSocialAdquiriente",
                        "fechaEmision",
                        "totalValorBruto",
                        "valorFactura",
                    ]
                )
                for it in (page.get("resultado") or []):
                    emisor = it.get("emisor") or {}
                    adquiriente = it.get("adquiriente") or {}
                    writer.writerow(
                        [
                            it.get("idFactura"),
                            it.get("numeroFactura"),
                            it.get("cufe"),
                            it.get("tipoFactura"),
                            emisor.get("nitEmisor"),
                            emisor.get("razonSocial"),
                            adquiriente.get("nitAdquiriente"),
                            adquiriente.get("razonSocial"),
                            it.get("fechaEmision"),
                            it.get("totalValorBruto"),
                            it.get("valorFactura"),
                        ]
                    )
                filename = f"siifa_facturas_{q['NitEmisor']}_p{q['NumeroPagina']}.csv"
                headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
                return Response(fp.getvalue(), mimetype="text/csv; charset=utf-8", headers=headers)
            if descargar == "xlsx":
                headers = [
                    "idFactura",
                    "numeroFactura",
                    "cufe",
                    "tipoFactura",
                    "nitEmisor",
                    "razonSocialEmisor",
                    "nitAdquiriente",
                    "razonSocialAdquiriente",
                    "fechaEmision",
                    "totalValorBruto",
                    "valorFactura",
                ]
                rows = []
                for it in (page.get("resultado") or []):
                    emisor = it.get("emisor") or {}
                    adquiriente = it.get("adquiriente") or {}
                    rows.append(
                        {
                            "idFactura": it.get("idFactura"),
                            "numeroFactura": it.get("numeroFactura"),
                            "cufe": it.get("cufe"),
                            "tipoFactura": it.get("tipoFactura"),
                            "nitEmisor": emisor.get("nitEmisor"),
                            "razonSocialEmisor": emisor.get("razonSocial"),
                            "nitAdquiriente": adquiriente.get("nitAdquiriente"),
                            "razonSocialAdquiriente": adquiriente.get("razonSocial"),
                            "fechaEmision": it.get("fechaEmision"),
                            "totalValorBruto": it.get("totalValorBruto"),
                            "valorFactura": it.get("valorFactura"),
                        }
                    )
                data = _xlsx_bytes("Facturas", headers, rows)
                filename = f"siifa_facturas_{q['NitEmisor']}_p{q['NumeroPagina']}.xlsx"
                resp_headers = {"Content-Disposition": f'attachment; filename=\"{filename}\"'}
                return Response(
                    data,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=resp_headers,
                )
        except SiifaApiError as e:
            error = str(e)
            details = json.dumps(e.payload, ensure_ascii=False, indent=2) if e.payload is not None else None
        except Exception as e:
            error = str(e)

    # construir prev/next query
    prev_q = dict(q)
    prev_q["NumeroPagina"] = max(1, int(q["NumeroPagina"]) - 1)
    next_q = dict(q)
    next_q["NumeroPagina"] = int(q["NumeroPagina"]) + 1

    html = HTML
    from jinja2 import Template
    tmpl = Template(html)
    body = tmpl.render(
        nav=_render_nav("facturas"),
        base_css=BASE_CSS,
        footer=_render_footer(),
        q=q,
        page=page,
        error=error,
        details=details,
        prev_q=urlencode({k: v for k, v in prev_q.items() if v not in (None, "")}),
        next_q=urlencode({k: v for k, v in next_q.items() if v not in (None, "")}),
    )
    return Response(body, mimetype="text/html; charset=utf-8")


SEGUIMIENTO_HTML = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Seguimiento (Glosas/Dev)</title>
  <style>{{ base_css|safe }}</style>
</head>
<body>
  {{ nav|safe }}
  <main class="container">
  <div class="page-title">
    <h1>Seguimiento (Glosas/Dev)</h1>
    <div class="meta">Consulta y exportación.</div>
  </div>
  {% if error %}
    <div class="error">
      <strong>Error:</strong> {{ error }}
      {% if details %}<pre>{{ details }}</pre>{% endif %}
    </div>
  {% endif %}
  <div class="card" style="margin-bottom: 14px;">
    <form method="get">
      <label>ID Factura<input name="IdFactura" value="{{ q.IdFactura or '' }}" /></label>
      <label>Número Factura<input name="NumeroFactura" value="{{ q.NumeroFactura or '' }}" /></label>
      <label>Nit Emisor<input name="NitEmisor" value="{{ q.NitEmisor or '' }}" /></label>
      <label>Tipo Seguimiento
        <select name="TipoSeguimiento">
          <option value="" {% if not q.TipoSeguimiento %}selected{% endif %}>Todos</option>
          <option value="GLOSA" {% if q.TipoSeguimiento == 'GLOSA' %}selected{% endif %}>Glosa</option>
          <option value="DEVOLUCION" {% if q.TipoSeguimiento == 'DEVOLUCION' %}selected{% endif %}>Devolución</option>
        </select>
      </label>
      <label>Tiene Respuesta
        <select name="TieneRespuesta">
          <option value="" {% if q.TieneRespuesta == None %}selected{% endif %}>Cualquiera</option>
          <option value="true" {% if q.TieneRespuesta == 'true' %}selected{% endif %}>Con respuesta</option>
          <option value="false" {% if q.TieneRespuesta == 'false' %}selected{% endif %}>Sin respuesta</option>
        </select>
      </label>
      <label>Registros por página<input name="RegistrosPorPagina" type="number" min="1" max="1500" value="{{ q.RegistrosPorPagina or 50 }}" /></label>
      <label>Número página<input name="NumeroPagina" type="number" min="1" value="{{ q.NumeroPagina or 1 }}" /></label>
      <div class="actions">
        <button type="submit">Buscar</button>
        <button type="submit" name="descargar" value="json">Descargar JSON</button>
        <button type="submit" name="descargar" value="xlsx">Descargar Excel</button>
        <a href="/responder-glosas{% if responder_link_q %}?{{ responder_link_q }}{% endif %}">Responder glosas</a>
        <a href="/responder-glosas{% if responder_download_q %}?{{ responder_download_q }}{% endif %}">Descargar pendientes</a>
      </div>
    </form>
  </div>

  {% if page %}
    <div class="meta">
      Total registros: {{ page.totalRegistros if page.totalRegistros is not none else 'N/D' }} |
      Total páginas: {{ page.totalPaginas if page.totalPaginas is not none else 'N/D' }} |
      Página actual: {{ page.paginaActual or q.NumeroPagina or 1 }}
    </div>
    <div class="pager">
      <a href="/seguimiento?{{ prev_q }}">« Anterior</a>
      <a href="/seguimiento?{{ next_q }}">Siguiente »</a>
    </div>
    <div class="grid">
      <table>
        <thead>
          <tr>
            <th>ID Seguimiento</th>
            <th>Tipo</th>
            <th>ID Factura</th>
            <th>Nro Factura</th>
            <th>Fecha Reporte</th>
            <th>Código</th>
            <th>Motivo</th>
            <th>Observación</th>
            <th>Respuesta</th>
          </tr>
        </thead>
        <tbody>
          {% for it in page.resultado or [] %}
          <tr>
            <td>{{ it.idSeguimientoFactura }}</td>
            <td>{{ it.tipoSeguimiento }}</td>
            <td>{{ it.idFactura }}</td>
            <td>{{ it.numeroFactura }}</td>
            <td>{{ it.fechaReporte }}</td>
            <td>{{ it.idSeguimientoTipoCodigo }}</td>
            <td>{{ it.descripcionSeguimientoTipoCodigo }}</td>
            <td style="max-width: 520px;">{{ it.observacion }}</td>
            <td style="max-width: 520px;">{{ it.descripcionSeguimientoTipoCodigoRespuesta }}</td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
  {% endif %}
  {{ footer|safe }}
  </main>
</body>
</html>
"""


@app.route("/seguimiento", methods=["GET"])
def seguimiento():
    token = _require_token()
    if not token:
        return redirect("/login")

    args = request.args
    q = {
        "IdFactura": args.get("IdFactura") or None,
        "NumeroFactura": args.get("NumeroFactura") or None,
        "NitEmisor": args.get("NitEmisor") or None,
        "TipoSeguimiento": args.get("TipoSeguimiento") or None,
        "TieneRespuesta": args.get("TieneRespuesta") or None,
        "RegistrosPorPagina": min(1500, max(1, int(args.get("RegistrosPorPagina") or 50))),
        "NumeroPagina": int(args.get("NumeroPagina") or 1),
    }

    page = None
    error = None
    details = None

    if any(v for k, v in q.items() if k not in ("RegistrosPorPagina", "NumeroPagina")) or args:
        try:
            client = _make_client_with_token(token)
            tiene = q["TieneRespuesta"]
            tiene_bool = None
            if tiene == "true":
                tiene_bool = True
            elif tiene == "false":
                tiene_bool = False

            page = client.list_seguimiento_factura(
                IdFactura=q["IdFactura"],
                NumeroFactura=q["NumeroFactura"],
                NitEmisor=q["NitEmisor"],
                TipoSeguimiento=q["TipoSeguimiento"],
                TieneRespuesta=tiene_bool,
                NumeroPagina=q["NumeroPagina"],
                RegistrosPorPagina=q["RegistrosPorPagina"],
            )
            factura_cache: dict[int, dict] = {}
            missing_ids: set[int] = set()
            for it in (page.get("resultado") or []):
                if not isinstance(it, dict):
                    continue
                if not it.get("numeroFactura") and q.get("NumeroFactura"):
                    it["numeroFactura"] = q.get("NumeroFactura")
                if not it.get("numeroFactura") and isinstance(it.get("facturaInfo"), dict):
                    it["numeroFactura"] = (it.get("facturaInfo") or {}).get("numeroFactura") or it.get("numeroFactura")
                if not it.get("numeroFactura") and it.get("idFactura"):
                    try:
                        fid = int(it.get("idFactura"))
                    except Exception:
                        fid = None
                    if fid:
                        if fid not in factura_cache:
                            factura_cache[fid] = _factura_by_id(client, fid)
                        it["numeroFactura"] = it.get("numeroFactura") or (factura_cache.get(fid) or {}).get("numeroFactura")
                        if not it.get("numeroFactura"):
                            missing_ids.add(fid)

            if missing_ids:
                extra = _facturas_map_by_ids(client, missing_ids)
                for it in (page.get("resultado") or []):
                    if not isinstance(it, dict):
                        continue
                    if it.get("numeroFactura"):
                        continue
                    try:
                        fid = int(it.get("idFactura"))
                    except Exception:
                        fid = None
                    if fid and fid in extra:
                        it["numeroFactura"] = extra[fid].get("numeroFactura") or it.get("numeroFactura")

            descargar = args.get("descargar")
            if descargar == "json":
                return Response(json.dumps(page, ensure_ascii=False, indent=2), mimetype="application/json")
            if descargar == "xlsx":
                headers = [
                    "idSeguimientoFactura",
                    "tipoSeguimiento",
                    "idFactura",
                    "numeroFactura",
                    "fechaReporte",
                    "idSeguimientoTipoCodigo",
                    "descripcionSeguimientoTipoCodigo",
                    "observacion",
                    "descripcionSeguimientoTipoCodigoRespuesta",
                ]
                rows = []
                factura_cache: dict[int, dict] = {}
                missing_ids: set[int] = set()
                for it in (page.get("resultado") or []):
                    if isinstance(it, dict):
                        row = {k: it.get(k) for k in headers}
                        if not row.get("numeroFactura"):
                            factura_info = it.get("facturaInfo") if isinstance(it.get("facturaInfo"), dict) else {}
                            row["numeroFactura"] = row.get("numeroFactura") or factura_info.get("numeroFactura")
                        if not row.get("numeroFactura") and it.get("idFactura"):
                            try:
                                fid = int(it.get("idFactura"))
                            except Exception:
                                fid = None
                            if fid:
                                if fid not in factura_cache:
                                    factura_cache[fid] = _factura_by_id(client, fid)
                                row["numeroFactura"] = row.get("numeroFactura") or (factura_cache.get(fid) or {}).get("numeroFactura")
                                if not row.get("numeroFactura"):
                                    missing_ids.add(fid)
                        if not row.get("numeroFactura") and q.get("NumeroFactura"):
                            row["numeroFactura"] = q.get("NumeroFactura")
                        rows.append(row)
                if missing_ids:
                    extra = _facturas_map_by_ids(client, missing_ids)
                    for row in rows:
                        if row.get("numeroFactura"):
                            continue
                        try:
                            fid = int(row.get("idFactura"))
                        except Exception:
                            fid = None
                        if fid and fid in extra:
                            row["numeroFactura"] = extra[fid].get("numeroFactura") or row.get("numeroFactura")
                data = _xlsx_bytes("Seguimiento", headers, rows)
                filename = f"siifa_seguimiento_p{q['NumeroPagina']}.xlsx"
                resp_headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
                return Response(
                    data,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=resp_headers,
                )
        except SiifaApiError as e:
            error = str(e)
            details = json.dumps(e.payload, ensure_ascii=False, indent=2) if e.payload is not None else None
        except Exception as e:
            error = str(e)

    prev_q = dict(q)
    prev_q["NumeroPagina"] = max(1, int(q["NumeroPagina"]) - 1)
    next_q = dict(q)
    next_q["NumeroPagina"] = int(q["NumeroPagina"]) + 1

    responder_params = {
        "IdFactura": q.get("IdFactura"),
        "NumeroFactura": q.get("NumeroFactura"),
        "NitEmisor": q.get("NitEmisor"),
        "RegistrosPorPagina": 1500,
    }
    responder_link_q = urlencode({k: v for k, v in responder_params.items() if v not in (None, "")})
    responder_download_q = urlencode({**{k: v for k, v in responder_params.items() if v not in (None, "")}, "descargar": "xlsx"})

    from jinja2 import Template
    body = Template(SEGUIMIENTO_HTML).render(
        nav=_render_nav("seguimiento"),
        base_css=BASE_CSS,
        footer=_render_footer(),
        q=q,
        page=page,
        error=error,
        details=details,
        prev_q=urlencode({k: v for k, v in prev_q.items() if v not in (None, "")}),
        next_q=urlencode({k: v for k, v in next_q.items() if v not in (None, "")}),
        responder_link_q=responder_link_q,
        responder_download_q=responder_download_q,
    )
    return Response(body, mimetype="text/html; charset=utf-8")


PAGOS_HTML = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Pagos</title>
  <style>{{ base_css|safe }}</style>
</head>
<body>
  {{ nav|safe }}
  <main class="container">
  <div class="page-title">
    <h1>Pagos</h1>
    <div class="meta">Consulta y exportación.</div>
  </div>
  {% if error %}
    <div class="error">
      <strong>Error:</strong> {{ error }}
      {% if details %}<pre>{{ details }}</pre>{% endif %}
    </div>
  {% endif %}
  <div class="card" style="margin-bottom: 14px;">
    <form method="get">
      <label>ID Factura<input name="IdFactura" value="{{ q.IdFactura or '' }}" /></label>
      <label>Número Factura<input name="NumeroFactura" value="{{ q.NumeroFactura or '' }}" /></label>
      <label>Nit Emisor<input name="NitEmisor" value="{{ q.NitEmisor or '' }}" /></label>
      <label>Nit Adquiriente<input name="NitAdquiriente" value="{{ q.NitAdquiriente or '' }}" /></label>
      <label>Código Fuente<input name="CodigoFuente" value="{{ q.CodigoFuente or '' }}" /></label>
      <label>Código SubFuente<input name="CodigoSubFuente" value="{{ q.CodigoSubFuente or '' }}" /></label>
      <label>Referencia Bancaria<input name="ReferenciaBancaria" value="{{ q.ReferenciaBancaria or '' }}" /></label>
      <label>Fecha Pago Inicio<input name="FechaPagoInicio" type="date" value="{{ q.FechaPagoInicio or '' }}" /></label>
      <label>Fecha Pago Final<input name="FechaPagoFinal" type="date" value="{{ q.FechaPagoFinal or '' }}" /></label>
      <label>Valor mínimo<input name="ValorMinimo" value="{{ q.ValorMinimo or '' }}" /></label>
      <label>Valor máximo<input name="ValorMaximo" value="{{ q.ValorMaximo or '' }}" /></label>
      <label>Registros por página<input name="RegistrosPorPagina" type="number" min="1" max="1500" value="{{ q.RegistrosPorPagina or 50 }}" /></label>
      <label>Número página<input name="NumeroPagina" type="number" min="1" value="{{ q.NumeroPagina or 1 }}" /></label>
      <div class="actions">
        <button type="submit">Buscar</button>
        <button type="submit" name="descargar" value="json">Descargar JSON</button>
        <button type="submit" name="descargar" value="xlsx">Descargar Excel</button>
      </div>
    </form>
  </div>

  {% if page %}
    <div class="meta">
      Total registros: {{ page.totalRegistros if page.totalRegistros is not none else 'N/D' }} |
      Total páginas: {{ page.totalPaginas if page.totalPaginas is not none else 'N/D' }} |
      Página actual: {{ page.paginaActual or q.NumeroPagina or 1 }}
    </div>
    <div class="pager">
      <a href="/pagos?{{ prev_q }}">« Anterior</a>
      <a href="/pagos?{{ next_q }}">Siguiente »</a>
    </div>
    <div class="grid">
      <table>
        <thead>
          <tr>
            <th>ID Pago</th>
            <th>ID Factura</th>
            <th>Nro Factura</th>
            <th>Fecha</th>
            <th>Valor</th>
            <th>Referencia</th>
            <th>Código Fuente</th>
            <th>Código SubFuente</th>
          </tr>
        </thead>
        <tbody>
          {% for it in page.resultado or [] %}
          <tr>
            <td>{{ it.idSeguimientoFacturaPago }}</td>
            <td>{{ it.idFactura }}</td>
            <td>{{ it.numeroFactura }}</td>
            <td>{{ it.fechaPago }}</td>
            <td>{{ it.valor }}</td>
            <td>{{ it.referenciaBancaria }}</td>
            <td>{{ it.codigoFuente }}</td>
            <td>{{ it.codigoSubFuente }}</td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
  {% endif %}
  {{ footer|safe }}
  </main>
</body>
</html>
"""


@app.route("/pagos", methods=["GET"])
def pagos():
    token = _require_token()
    if not token:
        return redirect("/login")

    args = request.args
    q = {
        "IdFactura": args.get("IdFactura") or None,
        "NumeroFactura": args.get("NumeroFactura") or None,
        "NitEmisor": args.get("NitEmisor") or None,
        "NitAdquiriente": args.get("NitAdquiriente") or None,
        "CodigoFuente": args.get("CodigoFuente") or None,
        "CodigoSubFuente": args.get("CodigoSubFuente") or None,
        "ReferenciaBancaria": args.get("ReferenciaBancaria") or None,
        "FechaPagoInicio": args.get("FechaPagoInicio") or None,
        "FechaPagoFinal": args.get("FechaPagoFinal") or None,
        "ValorMinimo": args.get("ValorMinimo") or None,
        "ValorMaximo": args.get("ValorMaximo") or None,
        "RegistrosPorPagina": min(1500, max(1, int(args.get("RegistrosPorPagina") or 50))),
        "NumeroPagina": int(args.get("NumeroPagina") or 1),
    }

    page = None
    error = None
    details = None

    if any(v for k, v in q.items() if k not in ("RegistrosPorPagina", "NumeroPagina")) or args:
        try:
            client = _make_client_with_token(token)
            page = client.list_pagos(**{k: v for k, v in q.items() if v not in (None, "")})

            descargar = args.get("descargar")
            if descargar == "json":
                return Response(json.dumps(page, ensure_ascii=False, indent=2), mimetype="application/json")
            if descargar == "xlsx":
                headers = [
                    "idSeguimientoFacturaPago",
                    "idFactura",
                    "numeroFactura",
                    "fechaPago",
                    "valor",
                    "referenciaBancaria",
                    "codigoFuente",
                    "codigoSubFuente",
                ]
                rows = []
                for it in (page.get("resultado") or []):
                    if isinstance(it, dict):
                        rows.append({k: it.get(k) for k in headers})
                data = _xlsx_bytes("Pagos", headers, rows)
                filename = f"siifa_pagos_p{q['NumeroPagina']}.xlsx"
                resp_headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
                return Response(
                    data,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=resp_headers,
                )
        except SiifaApiError as e:
            error = str(e)
            details = json.dumps(e.payload, ensure_ascii=False, indent=2) if e.payload is not None else None
        except Exception as e:
            error = str(e)

    prev_q = dict(q)
    prev_q["NumeroPagina"] = max(1, int(q["NumeroPagina"]) - 1)
    next_q = dict(q)
    next_q["NumeroPagina"] = int(q["NumeroPagina"]) + 1

    from jinja2 import Template
    body = Template(PAGOS_HTML).render(
        nav=_render_nav("pagos"),
        base_css=BASE_CSS,
        footer=_render_footer(),
        q=q,
        page=page,
        error=error,
        details=details,
        prev_q=urlencode({k: v for k, v in prev_q.items() if v not in (None, "")}),
        next_q=urlencode({k: v for k, v in next_q.items() if v not in (None, "")}),
    )
    return Response(body, mimetype="text/html; charset=utf-8")


CONSULTA_MASIVA_HTML = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Consultas masivas</title>
  <style>{{ base_css|safe }}</style>
</head>
<body>
  {{ nav|safe }}
  <main class="container">
  <div class="page-title">
    <h1>Consultas masivas</h1>
    <div class="meta">Sube Excel/CSV/JSON con idFactura o numeroFactura + nitEmisor.</div>
  </div>
  {% if error %}
    <div class="error">
      <strong>Error:</strong> {{ error }}
      {% if details %}<pre>{{ details }}</pre>{% endif %}
    </div>
  {% endif %}
  <div class="card">
    <form method="post" enctype="multipart/form-data">
      <input type="file" name="archivo" accept=".xlsx,.csv,.json,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,text/csv,application/json" required />
      <div class="actions">
        <button type="submit">Procesar</button>
        {% if last_available %}
          <a href="/consulta-masiva/descargar?fmt=xlsx">Descargar último resultado (Excel)</a>
          <a href="/consulta-masiva/descargar?fmt=csv">Descargar último resultado (CSV)</a>
        {% endif %}
      </div>
    </form>
    {% if rows %}
      <div style="margin-top:12px;" class="meta">Vista previa ({{ rows|length }} de {{ total_rows }} filas)</div>
      <div class="grid" style="margin-top:10px;">
        <table>
          <thead>
            <tr>
              <th>idFactura</th>
              <th>numeroFactura</th>
              <th>nitEmisor</th>
              <th>nitAdquiriente</th>
              <th>tieneRadicado</th>
              <th>tieneGlosa</th>
              <th>tieneDevolucion</th>
              <th>tienePagos</th>
              <th>totalPagos</th>
              <th>warning</th>
              <th>error</th>
            </tr>
          </thead>
          <tbody>
            {% for it in rows %}
            <tr>
              <td>{{ it.idFactura }}</td>
              <td>{{ it.numeroFactura }}</td>
              <td>{{ it.nitEmisor }}</td>
              <td>{{ it.nitAdquiriente }}</td>
              <td>{{ it.tieneRadicado }}</td>
              <td>{{ it.tieneGlosa }}</td>
              <td>{{ it.tieneDevolucion }}</td>
              <td>{{ it.tienePagos }}</td>
              <td>{{ it.totalPagos }}</td>
              <td>{{ it.warning }}</td>
              <td>{{ it.error }}</td>
            </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    {% endif %}
  </div>
  {{ footer|safe }}
  </main>
</body>
</html>
"""


def _estado_factura(client: SiifaClient, id_factura: int) -> dict:
    factura = client.get_factura(id_factura)
    numero_factura = factura.get("numeroFactura")
    emisor = factura.get("emisor") or {}
    adquiriente = factura.get("adquiriente") or {}

    radicados = []
    try:
        radicados = client.list_radicados_by_id_factura(id_factura)
    except Exception:
        radicados = []

    seguimiento = client.list_seguimiento_factura(IdFactura=id_factura, NumeroPagina=1, RegistrosPorPagina=1500)
    seg_items = seguimiento.get("resultado") or []
    tipos = {(it.get("tipoSeguimiento") or "").upper() for it in seg_items if isinstance(it, dict)}
    tiene_glosa = "GLOSA" in tipos
    tiene_devolucion = "DEVOLUCION" in tipos or "DEVOLUCIÓN" in tipos

    pagos = client.list_seguimiento_pago(IdFactura=id_factura, NumeroPagina=1, RegistrosPorPagina=1500)
    pagos_items = pagos.get("resultado") or []
    total_pagos = 0.0
    for it in pagos_items:
        if isinstance(it, dict):
            v = _coerce_float(it.get("valor"))
            if v:
                total_pagos += float(v)

    return {
        "idFactura": id_factura,
        "numeroFactura": numero_factura,
        "nitEmisor": emisor.get("nitEmisor"),
        "nitAdquiriente": adquiriente.get("nitAdquiriente"),
        "tieneRadicado": "SI" if radicados else "NO",
        "tieneGlosa": "SI" if tiene_glosa else "NO",
        "tieneDevolucion": "SI" if tiene_devolucion else "NO",
        "tienePagos": "SI" if pagos_items else "NO",
        "totalPagos": float(total_pagos),
    }


@app.route("/consulta-masiva", methods=["GET", "POST"])
def consulta_masiva():
    token = _require_token()
    if not token:
        return redirect("/login")

    error = None
    details = None
    rows_preview = None
    total_rows = 0

    if request.method == "POST":
        try:
            file = request.files.get("archivo")
            if not file:
                raise ValueError("Debe adjuntar un archivo Excel (.xlsx), CSV o JSON")
            filename = (file.filename or "").lower()
            if filename.endswith(".json"):
                payload = _parse_uploaded_json(file)
                if isinstance(payload, dict):
                    for key in ("items", "facturas", "lista", "listaFacturas"):
                        if key in payload and isinstance(payload[key], list):
                            payload = payload[key]
                            break
                if not isinstance(payload, list):
                    raise ValueError("JSON inválido. Debe ser una lista o un objeto con una lista (items/facturas/lista).")
                in_rows = payload
            else:
                if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
                    in_rows = _parse_uploaded_xlsx(file)
                else:
                    in_rows = _parse_uploaded_csv(file)

            client = _make_client_with_token(token)
            out_rows = []
            for r in in_rows:
                total_rows += 1
                warning = None
                id_factura = None
                try:
                    if isinstance(r, (int, float, str)):
                        id_factura = _coerce_int(r)
                        meta = {}
                    elif isinstance(r, dict):
                        id_factura, meta = _resolve_id_factura(client, r)
                    else:
                        id_factura, meta = None, {"warning": "Fila inválida"}
                    warning = meta.get("warning")
                    if not id_factura:
                        out = {
                            "idFactura": None,
                            "numeroFactura": None,
                            "nitEmisor": None,
                            "nitAdquiriente": None,
                            "tieneRadicado": "NO",
                            "tieneGlosa": "NO",
                            "tieneDevolucion": "NO",
                            "tienePagos": "NO",
                            "totalPagos": 0.0,
                            "warning": warning,
                            "error": None,
                        }
                    else:
                        out = _estado_factura(client, int(id_factura))
                        out["warning"] = warning
                        out["error"] = None
                except SiifaApiError as e:
                    out = {
                        "idFactura": id_factura,
                        "numeroFactura": None,
                        "nitEmisor": None,
                        "nitAdquiriente": None,
                        "tieneRadicado": "NO",
                        "tieneGlosa": "NO",
                        "tieneDevolucion": "NO",
                        "tienePagos": "NO",
                        "totalPagos": 0.0,
                        "warning": warning,
                        "error": str(e),
                    }
                out_rows.append(out)

            fp = io.StringIO()
            writer = csv.writer(fp)
            writer.writerow(
                [
                    "idFactura",
                    "numeroFactura",
                    "nitEmisor",
                    "nitAdquiriente",
                    "tieneRadicado",
                    "tieneGlosa",
                    "tieneDevolucion",
                    "tienePagos",
                    "totalPagos",
                    "warning",
                    "error",
                ]
            )
            for it in out_rows:
                writer.writerow(
                    [
                        it.get("idFactura"),
                        it.get("numeroFactura"),
                        it.get("nitEmisor"),
                        it.get("nitAdquiriente"),
                        it.get("tieneRadicado"),
                        it.get("tieneGlosa"),
                        it.get("tieneDevolucion"),
                        it.get("tienePagos"),
                        it.get("totalPagos"),
                        it.get("warning"),
                        it.get("error"),
                    ]
                )
            _set_session_value("consulta_masiva_last_csv", fp.getvalue())
            _set_session_value("consulta_masiva_last_obj", out_rows)
            rows_preview = out_rows[:50]
        except SiifaApiError as e:
            error = str(e)
            details = json.dumps(e.payload, ensure_ascii=False, indent=2) if e.payload is not None else None
        except Exception as e:
            error = str(e)

    last_available = _get_session_value("consulta_masiva_last_csv") is not None or _get_session_value("consulta_masiva_last_obj") is not None
    from jinja2 import Template
    body = Template(CONSULTA_MASIVA_HTML).render(
        nav=_render_nav("consulta"),
        base_css=BASE_CSS,
        footer=_render_footer(),
        error=error,
        details=details,
        rows=rows_preview,
        total_rows=total_rows,
        last_available=last_available,
    )
    return Response(body, mimetype="text/html; charset=utf-8")


@app.route("/consulta-masiva/descargar", methods=["GET"])
def consulta_masiva_descargar():
    token = _require_token()
    if not token:
        return redirect("/login")
    fmt = (request.args.get("fmt") or "csv").strip().lower()
    text = _get_session_value("consulta_masiva_last_csv")
    last_obj = _get_session_value("consulta_masiva_last_obj")
    if not text and last_obj is None:
        return redirect("/consulta-masiva")
    if fmt == "xlsx":
        headers = [
            "idFactura",
            "numeroFactura",
            "nitEmisor",
            "nitAdquiriente",
            "tieneRadicado",
            "tieneGlosa",
            "tieneDevolucion",
            "tienePagos",
            "totalPagos",
            "warning",
            "error",
        ]
        rows = [it for it in (last_obj or []) if isinstance(it, dict)]
        data = _xlsx_bytes("ConsultaMasiva", headers, rows)
        resp_headers = {"Content-Disposition": 'attachment; filename="siifa_consulta_masiva_ultimo.xlsx"'}
        return Response(
            data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=resp_headers,
        )
    headers = {"Content-Disposition": 'attachment; filename="siifa_consulta_masiva_ultimo.csv"'}
    return Response(text or "", mimetype="text/csv; charset=utf-8", headers=headers)


CARGUE_HOME_HTML = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Cargue masivo</title>
  <style>{{ base_css|safe }}</style>
</head>
<body>
  {{ nav|safe }}
  <main class="container">
  <div class="page-title">
    <h1>Cargue masivo</h1>
    <div class="meta">Carga de archivos para operaciones masivas.</div>
  </div>
  <div class="cards">
    <div class="card">
      <h3>Glosas</h3>
      <p>POST /api/SeguimientoFacturaGlosa/Masivo (si 403: modo asistido)</p>
      <div class="actions">
        <a href="/cargue/glosas">Abrir</a>
        <a href="/cargue/glosas/plantilla.xlsx">Plantilla</a>
      </div>
    </div>
  </div>
  {{ footer|safe }}
  </main>
</body>
</html>
"""


@app.route("/cargue", methods=["GET"])
def cargue_home():
    token = _require_token()
    if not token:
        return redirect("/login")
    from jinja2 import Template
    body = Template(CARGUE_HOME_HTML).render(nav=_render_nav("cargue"), base_css=BASE_CSS, footer=_render_footer())
    return Response(body, mimetype="text/html; charset=utf-8")


RESPONDER_GLOSAS_HTML = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Responder glosas</title>
  <style>{{ base_css|safe }}</style>
</head>
<body>
  {{ nav|safe }}
  <main class="container">
  <div class="page-title">
    <h1>Responder glosas</h1>
    <div class="meta">Descarga pendientes, completa respuestas y sube el Excel.</div>
  </div>

  {% if error %}
    <div class="error">
      <strong>Error:</strong> {{ error }}
      {% if details %}<pre>{{ details }}</pre>{% endif %}
    </div>
  {% endif %}

  <div class="card" style="margin-bottom: 14px;">
    <form method="get">
      <label>ID Factura<input name="IdFactura" value="{{ q.IdFactura or '' }}" /></label>
      <label>Número Factura<input name="NumeroFactura" value="{{ q.NumeroFactura or '' }}" /></label>
      <label>Nit Emisor<input name="NitEmisor" value="{{ q.NitEmisor or '' }}" /></label>
      <label>Registros por página<input name="RegistrosPorPagina" type="number" min="1" max="1500" value="{{ q.RegistrosPorPagina or 1500 }}" /></label>
      <div class="actions">
        <button type="submit">Buscar</button>
        <button type="submit" name="descargar" value="xlsx">Descargar pendientes (Excel)</button>
      </div>
    </form>
  </div>

  {% if pendientes %}
    <div class="meta">Pendientes (vista previa {{ pendientes|length }}{% if pendientes_total %} de {{ pendientes_total }}{% endif %})</div>
    <div class="grid" style="margin-top:10px; margin-bottom: 14px;">
      <table>
        <thead>
          <tr>
            <th>idSeguimientoFacturaGlosa</th>
            <th>idFactura</th>
            <th>numeroFactura</th>
            <th>nitEmisor</th>
            <th>idSeguimientoTipoCodigoGlosa</th>
            <th>fechaFormulacion</th>
            <th>valorGlosa</th>
            <th>observacion</th>
          </tr>
        </thead>
        <tbody>
          {% for it in pendientes %}
          <tr>
            <td>{{ it.idSeguimientoFacturaGlosa }}</td>
            <td>{{ it.idFactura }}</td>
            <td>{{ it.numeroFactura }}</td>
            <td>{{ it.nitEmisor }}</td>
            <td>{{ it.idSeguimientoTipoCodigoGlosa }}</td>
            <td>{{ it.fechaFormulacion }}</td>
            <td>{{ it.valorGlosa }}</td>
            <td style="max-width:520px;">{{ it.observacion }}</td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
  {% endif %}

  <div class="card" style="margin-bottom: 14px;">
    <form method="post" enctype="multipart/form-data">
      <div class="meta" style="margin-bottom:8px;">Sube el Excel ya diligenciado con columnas de respuesta.</div>
      <input type="file" name="archivo" accept=".xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" required />
      <div class="actions">
        <button type="submit">Cargar respuestas</button>
        {% if last_available %}
          <a href="/responder-glosas/descargar?fmt=xlsx">Descargar último resultado (Excel)</a>
          <a href="/responder-glosas/descargar?fmt=json">Descargar último resultado (JSON)</a>
        {% endif %}
      </div>
    </form>
  </div>

  {% if preview %}
    <div class="meta">Vista previa ({{ preview|length }} filas)</div>
    <div class="grid" style="margin-top:10px;">
      <table>
        <thead>
          <tr>
            <th>idSeguimientoFacturaGlosa</th>
            <th>numeroFactura</th>
            <th>idSeguimientoTipoCodigoGlosa</th>
            <th>valorGlosa</th>
            <th>observacion</th>
            <th>idSeguimientoTipoCodigoRespuesta</th>
            <th>fechaRespuesta</th>
            <th>observacionRespuesta</th>
            <th>ok</th>
            <th>error</th>
          </tr>
        </thead>
        <tbody>
        {% for it in preview %}
          <tr>
            <td>{{ it.idSeguimientoFacturaGlosa }}</td>
            <td>{{ it.numeroFactura }}</td>
            <td>{{ it.idSeguimientoTipoCodigoGlosa }}</td>
            <td>{{ it.valorGlosa }}</td>
            <td style="max-width:520px;">{{ it.observacion }}</td>
            <td>{{ it.idSeguimientoTipoCodigoRespuesta }}</td>
            <td>{{ it.fechaRespuesta }}</td>
            <td style="max-width:520px;">{{ it.observacionRespuesta }}</td>
            <td>{{ it.ok }}</td>
            <td style="max-width:520px;">{{ it.error }}</td>
          </tr>
        {% endfor %}
        </tbody>
      </table>
    </div>
  {% endif %}
  {{ footer|safe }}
  </main>
</body>
</html>
"""


@app.route("/responder-glosas", methods=["GET", "POST"])
def responder_glosas():
    token = _require_token()
    if not token:
        return redirect("/login")

    error = None
    details = None
    preview = None
    pendientes_preview = None
    pendientes_total = None

    if request.method == "POST":
        try:
            file = request.files.get("archivo")
            if not file:
                raise ValueError("Debe adjuntar un archivo Excel (.xlsx)")
            in_rows = _parse_uploaded_xlsx(file)

            client = _make_client_with_token(token)
            resultados = []
            for r in in_rows:
                if not isinstance(r, dict):
                    continue
                sid = _coerce_int(_row_get(r, "idSeguimientoFacturaGlosa", "IdSeguimientoFacturaGlosa", "idSeguimientoFactura"))
                if sid is None or int(sid) <= 0:
                    continue
                resp_code = (str(_row_get(r, "idSeguimientoTipoCodigoRespuesta", "IdSeguimientoTipoCodigoRespuesta") or "").strip() or None)
                fecha_resp = _to_iso_z(_row_get(r, "fechaRespuesta", "FechaRespuesta"))
                obs_resp = (str(_row_get(r, "observacionRespuesta", "ObservacionRespuesta") or "").strip() or None)

                numero_factura = (str(_row_get(r, "numeroFactura", "NumeroFactura") or "").strip() or None)
                tipo_glosa = (str(_row_get(r, "idSeguimientoTipoCodigoGlosa", "IdSeguimientoTipoCodigoGlosa") or "").strip() or None)
                valor_glosa = _coerce_float(_row_get(r, "valorGlosa", "ValorGlosa"))
                obs = (str(_row_get(r, "observacion", "Observacion") or "").strip() or None)

                item_out = {
                    "idSeguimientoFacturaGlosa": sid,
                    "numeroFactura": numero_factura,
                    "idSeguimientoTipoCodigoGlosa": tipo_glosa,
                    "valorGlosa": valor_glosa,
                    "observacion": obs,
                    "idSeguimientoTipoCodigoRespuesta": resp_code,
                    "fechaRespuesta": fecha_resp,
                    "observacionRespuesta": obs_resp,
                }

                try:
                    if not sid or not resp_code or not fecha_resp:
                        raise ValueError("Faltan campos obligatorios: idSeguimientoFacturaGlosa, idSeguimientoTipoCodigoRespuesta, fechaRespuesta")
                    if obs_resp is not None and len(str(obs_resp)) > 450:
                        raise ValueError("observacionRespuesta supera 450 caracteres. Redúzcala o resuma el texto.")
                    payload = {
                        "idSeguimientoFacturaGlosa": int(sid),
                        "idSeguimientoTipoCodigoRespuesta": resp_code,
                        "fechaRespuesta": fecha_resp,
                        "observacionRespuesta": obs_resp,
                    }
                    client.responder_glosa(payload)
                    item_out["ok"] = True
                    item_out["error"] = None
                except SiifaApiError as e:
                    item_out["ok"] = False
                    item_out["error"] = str(e)
                    item_out["payload"] = e.payload
                except Exception as e:
                    item_out["ok"] = False
                    item_out["error"] = str(e)
                    item_out["payload"] = None

                resultados.append(item_out)

            _set_session_value("responder_glosas_last_obj", resultados)
            _set_session_value("responder_glosas_last_json", json.dumps(resultados, ensure_ascii=False, indent=2))
            preview = resultados[:50]
        except SiifaApiError as e:
            error = str(e)
            details = json.dumps(e.payload, ensure_ascii=False, indent=2) if e.payload is not None else None
        except Exception as e:
            error = str(e)

    args = request.args
    q = {
        "IdFactura": args.get("IdFactura") or None,
        "NumeroFactura": args.get("NumeroFactura") or None,
        "NitEmisor": args.get("NitEmisor") or None,
        "RegistrosPorPagina": min(1500, max(1, int(args.get("RegistrosPorPagina") or 1500))),
    }

    def _build_pendientes(client: SiifaClient):
        page = client.list_seguimiento_factura(
            IdFactura=q["IdFactura"],
            NumeroFactura=q["NumeroFactura"],
            NitEmisor=q["NitEmisor"],
            TipoSeguimiento="GLOSA",
            TieneRespuesta=False,
            NumeroPagina=1,
            RegistrosPorPagina=q["RegistrosPorPagina"],
        )
        items = page.get("resultado") or []
        out = []
        factura_cache: dict[int, dict] = {}
        missing_ids: set[int] = set()
        for it in items:
            if not isinstance(it, dict):
                continue
            factura_info = it.get("facturaInfo") if isinstance(it.get("facturaInfo"), dict) else {}
            emisor = (factura_info.get("emisor") or {}) if isinstance(factura_info.get("emisor"), dict) else {}
            sid = it.get("idSeguimientoFacturaGlosa") or it.get("idSeguimientoFactura")
            numero = it.get("numeroFactura") or factura_info.get("numeroFactura")
            nit_emisor = emisor.get("nitEmisor") or it.get("nitEmisor")
            if not numero and q.get("NumeroFactura"):
                numero = q.get("NumeroFactura")
            if not nit_emisor and q.get("NitEmisor"):
                nit_emisor = q.get("NitEmisor")
            if (not numero or not nit_emisor) and it.get("idFactura"):
                try:
                    fid = int(it.get("idFactura"))
                except Exception:
                    fid = None
                if fid:
                    if fid not in factura_cache:
                        factura_cache[fid] = _factura_by_id(client, fid)
                    fac = factura_cache.get(fid) or {}
                    if not numero:
                        numero = fac.get("numeroFactura") or numero
                    if not nit_emisor:
                        em = fac.get("emisor") if isinstance(fac.get("emisor"), dict) else {}
                        nit_emisor = (em or {}).get("nitEmisor") or nit_emisor
                    if (not numero or not nit_emisor):
                        missing_ids.add(fid)
            out.append(
                {
                    "idSeguimientoFacturaGlosa": sid,
                    "idFactura": it.get("idFactura"),
                    "numeroFactura": numero,
                    "nitEmisor": nit_emisor,
                    "idSeguimientoTipoCodigoGlosa": it.get("idSeguimientoTipoCodigo") or it.get("idSeguimientoTipoCodigoGlosa"),
                    "descripcionSeguimientoTipoCodigoGlosa": it.get("descripcionSeguimientoTipoCodigo"),
                    "fechaFormulacion": it.get("fechaFormulacion"),
                    "valorGlosa": it.get("valor"),
                    "observacion": it.get("observacion"),
                    "idSeguimientoTipoCodigoRespuesta": None,
                    "fechaRespuesta": None,
                    "observacionRespuesta": None,
                }
            )
        if missing_ids:
            extra = _facturas_map_by_ids(client, missing_ids)
            for r in out:
                if r.get("numeroFactura") and r.get("nitEmisor"):
                    continue
                try:
                    fid = int(r.get("idFactura"))
                except Exception:
                    fid = None
                if fid and fid in extra:
                    if not r.get("numeroFactura"):
                        r["numeroFactura"] = extra[fid].get("numeroFactura") or r.get("numeroFactura")
                    if not r.get("nitEmisor"):
                        r["nitEmisor"] = extra[fid].get("nitEmisor") or r.get("nitEmisor")
        return page, out

    if request.method == "GET" and args.get("descargar") == "xlsx":
        try:
            client = _make_client_with_token(token)
            page, pendientes = _build_pendientes(client)

            ejemplo = {
                "idSeguimientoFacturaGlosa": 0,
                "idFactura": 0,
                "numeroFactura": "EJEMPLO-FACTURA",
                "nitEmisor": "900000000",
                "idSeguimientoTipoCodigoGlosa": "CO2301",
                "descripcionSeguimientoTipoCodigoGlosa": "Ejemplo motivo",
                "fechaFormulacion": "2026-04-23T16:48:48Z",
                "valorGlosa": 17500,
                "observacion": "Ejemplo observación glosa",
                "idSeguimientoTipoCodigoRespuesta": "RESP01",
                "fechaRespuesta": "2026-04-24T10:00:00Z",
                "observacionRespuesta": "Ejemplo respuesta (texto)",
            }
            pendientes = [ejemplo] + pendientes

            catalogo = client.list_seguimiento_tipo_codigo_by_grupo("RESPUESTA")
            cat_rows = []
            if isinstance(catalogo, list):
                for c in catalogo:
                    if isinstance(c, dict):
                        cat_rows.append(
                            {
                                "idSeguimientoTipoCodigo": c.get("idSeguimientoTipoCodigo"),
                                "descripcion": c.get("descripcion"),
                                "nivel": c.get("nivel"),
                                "grupo": c.get("grupo"),
                                "activo": c.get("activo"),
                            }
                        )

            data = _xlsx_bytes_multi(
                [
                    (
                        "Pendientes",
                        [
                            "idSeguimientoFacturaGlosa",
                            "idFactura",
                            "numeroFactura",
                            "nitEmisor",
                            "idSeguimientoTipoCodigoGlosa",
                            "descripcionSeguimientoTipoCodigoGlosa",
                            "fechaFormulacion",
                            "valorGlosa",
                            "observacion",
                            "idSeguimientoTipoCodigoRespuesta",
                            "fechaRespuesta",
                            "observacionRespuesta",
                        ],
                        pendientes,
                    ),
                    (
                        "CatalogoRespuesta",
                        ["idSeguimientoTipoCodigo", "descripcion", "nivel", "grupo", "activo"],
                        cat_rows,
                    ),
                ]
            )
            filename = "siifa_glosas_pendientes_para_responder.xlsx"
            headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
            return Response(
                data,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers=headers,
            )
        except SiifaApiError as e:
            error = str(e)
            details = json.dumps(e.payload, ensure_ascii=False, indent=2) if e.payload is not None else None
        except Exception as e:
            error = str(e)
    elif request.method == "GET" and (args.get("IdFactura") or args.get("NumeroFactura") or args.get("NitEmisor")):
        try:
            client = _make_client_with_token(token)
            page, pendientes = _build_pendientes(client)
            pendientes_total = page.get("totalRegistros")
            pendientes_preview = pendientes[:50]
        except SiifaApiError as e:
            error = str(e)
            details = json.dumps(e.payload, ensure_ascii=False, indent=2) if e.payload is not None else None
        except Exception as e:
            error = str(e)

    last_available = _get_session_value("responder_glosas_last_json") is not None or _get_session_value("responder_glosas_last_obj") is not None
    from jinja2 import Template
    body = Template(RESPONDER_GLOSAS_HTML).render(
        nav=_render_nav("responder"),
        base_css=BASE_CSS,
        footer=_render_footer(),
        q=q,
        error=error,
        details=details,
        preview=preview,
        pendientes=pendientes_preview,
        pendientes_total=pendientes_total,
        last_available=last_available,
    )
    return Response(body, mimetype="text/html; charset=utf-8")


@app.route("/responder-glosas/descargar", methods=["GET"])
def responder_glosas_descargar():
    token = _require_token()
    if not token:
        return redirect("/login")
    fmt = (request.args.get("fmt") or "xlsx").strip().lower()
    text = _get_session_value("responder_glosas_last_json")
    last_obj = _get_session_value("responder_glosas_last_obj")
    if not text and last_obj is None:
        return redirect("/responder-glosas")
    if fmt == "json":
        headers = {"Content-Disposition": 'attachment; filename="siifa_respuesta_glosas_ultimo.json"'}
        return Response(text or "", mimetype="application/json; charset=utf-8", headers=headers)
    rows = [it for it in (last_obj or []) if isinstance(it, dict)]
    headers_cols = [
        "idSeguimientoFacturaGlosa",
        "numeroFactura",
        "idSeguimientoTipoCodigoGlosa",
        "valorGlosa",
        "observacion",
        "idSeguimientoTipoCodigoRespuesta",
        "fechaRespuesta",
        "observacionRespuesta",
        "ok",
        "error",
    ]
    data = _xlsx_bytes("Resultado", headers_cols, rows)
    headers = {"Content-Disposition": 'attachment; filename="siifa_respuesta_glosas_ultimo.xlsx"'}
    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


CARGUE_GLOSAS_HTML = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Cargue masivo de glosas</title>
  <style>{{ base_css|safe }}</style>
</head>
<body>
  {{ nav|safe }}
  <main class="container">
  <div class="page-title">
    <h1>Cargue masivo de glosas</h1>
    <div class="meta">Excel/CSV/JSON. Si masivo da 403, usa modo asistido (una por una).</div>
  </div>
  {% if error %}
    <div class="error">
      <strong>Error:</strong> {{ error }}
      {% if details %}<pre>{{ details }}</pre>{% endif %}
    </div>
  {% endif %}
  <form method="post" enctype="multipart/form-data">
    <div class="card">
      <div style="color:#555; font-size:13px; margin-bottom: 8px;">
        Suba un Excel (.xlsx), CSV o JSON.
      </div>
      <input type="file" name="archivo" accept=".xlsx,.csv,.json,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,text/csv,application/json" required />
      <div class="actions">
        <button type="submit">Procesar cargue</button>
        <a href="/cargue/glosas/plantilla.xlsx">Descargar plantilla (Excel)</a>
        {% if last_available %}
          <a href="/cargue/glosas/descargar?fmt=xlsx">Descargar última respuesta (Excel)</a>
          <a href="/cargue/glosas/descargar?fmt=json">Descargar última respuesta (JSON)</a>
        {% endif %}
      </div>
      {% if result %}
        <div style="margin-top: 12px; color:#555; font-size:13px;">Resultado</div>
        <pre>{{ result }}</pre>
      {% endif %}
    </div>
  </form>
  {{ footer|safe }}
  </main>
</body>
</html>
"""


def _z2(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return str(int(float(s))).zfill(2)
    except Exception:
        return s.zfill(2)


def _glosas_from_tabular(client: SiifaClient, in_rows: list[dict]) -> list[dict]:
    out = []
    for r in in_rows:
        if not isinstance(r, dict):
            continue
        id_factura = _coerce_int(_row_get(r, "idFactura", "IdFactura"))
        if not id_factura:
            resolved, _meta = _resolve_id_factura(client, r)
            id_factura = resolved

        id_tipo = (str(_row_get(r, "idSeguimientoTipoCodigoGlosa", "tipoGlosa", "TipoGlosa") or "").strip() or None)
        if not id_tipo:
            cg = (str(_row_get(r, "CODIGO CONCEPTO GENERAL", "codigoConceptoGeneral") or "").strip() or None)
            ce = _z2(_row_get(r, "CODIGO CONCEPTO ESPECIFICO", "codigoConceptoEspecifico"))
            ca = _z2(_row_get(r, "CODIGO CONCEPTO APLICACION", "codigoConceptoAplicacion"))
            if cg and ce and ca:
                id_tipo = f"{cg}{ce}{ca}"

        fecha = _to_iso_z(
            _row_get(
                r,
                "fechaFormulacion",
                "fecha_formulacion",
                "FECHA DE RADICACION",
                "FECHA DE FACTURA",
            )
        )
        valor = _coerce_float(_row_get(r, "valorGlosa", "valor_glosa", "VALOR GLOSADO GLOSA"))
        obs = (str(_row_get(r, "observacion", "Observacion", "OBSERVACIONES") or "").strip() or None)
        if not obs:
            obs = (str(_row_get(r, "VALIDACION DE LA GLOSA") or "").strip() or None)

        out.append(
            {
                "idFactura": id_factura,
                "idSeguimientoTipoCodigoGlosa": id_tipo,
                "fechaFormulacion": fecha,
                "valorGlosa": valor,
                "observacion": obs,
            }
        )
    return out


@app.route("/cargue/glosas", methods=["GET", "POST"])
def cargue_glosas():
    token = _require_token()
    if not token:
        return redirect("/login")

    error = None
    details = None
    result = None

    if request.method == "POST":
        try:
            file = request.files.get("archivo")
            if not file:
                raise ValueError("Debe adjuntar un archivo Excel (.xlsx), CSV o JSON")
            filename = (file.filename or "").lower()
            client = _make_client_with_token(token)

            if filename.endswith(".json"):
                payload = _parse_uploaded_json(file)
                if isinstance(payload, dict) and "listaGlosas" in payload:
                    lista = payload["listaGlosas"]
                else:
                    lista = payload
                if not isinstance(lista, list):
                    raise ValueError("Entrada inválida. Debe ser una lista o un objeto con listaGlosas.")
            else:
                if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
                    in_rows = _parse_uploaded_xlsx(file)
                else:
                    in_rows = _parse_uploaded_csv(file)
                lista = _glosas_from_tabular(client, in_rows)

            try:
                resp = client.crear_glosas_masivo(lista)
                modo = "masivo"
            except SiifaApiError as e:
                if e.status != 403:
                    raise
                modo = "asistido"
                resp = []
                for idx, it in enumerate(lista, start=1):
                    try:
                        if not it.get("idFactura"):
                            raise ValueError("No se pudo resolver idFactura")
                        resp.append({"index": idx, "ok": True, "response": client.crear_glosa(it)})
                    except SiifaApiError as ie:
                        resp.append({"index": idx, "ok": False, "error": str(ie), "payload": ie.payload})
                    except Exception as ie:
                        resp.append({"index": idx, "ok": False, "error": str(ie), "payload": None})

            result = json.dumps({"modo": modo, "resultado": resp}, ensure_ascii=False, indent=2)
            _set_session_value("cargue_glosas_last_json", result)
            _set_session_value("cargue_glosas_last_obj", {"modo": modo, "resultado": resp})
        except SiifaApiError as e:
            error = str(e)
            details = json.dumps(e.payload, ensure_ascii=False, indent=2) if e.payload is not None else None
        except Exception as e:
            error = str(e)

    last_available = _get_session_value("cargue_glosas_last_json") is not None or _get_session_value("cargue_glosas_last_obj") is not None
    from jinja2 import Template

    tmpl = Template(CARGUE_GLOSAS_HTML)
    body = tmpl.render(
        nav=_render_nav("cargue"),
        base_css=BASE_CSS,
        footer=_render_footer(),
        error=error,
        details=details,
        result=result,
        last_available=last_available,
    )
    return Response(body, mimetype="text/html; charset=utf-8")


@app.route("/cargue/glosas/descargar", methods=["GET"])
def cargue_glosas_descargar():
    token = _require_token()
    if not token:
        return redirect("/login")
    fmt = (request.args.get("fmt") or "json").strip().lower()
    text = _get_session_value("cargue_glosas_last_json")
    last_obj = _get_session_value("cargue_glosas_last_obj")
    if not text and last_obj is None:
        return redirect("/cargue/glosas")
    if fmt == "xlsx":
        rows = []
        if isinstance(last_obj, dict) and isinstance(last_obj.get("resultado"), list):
            for it in last_obj["resultado"]:
                if isinstance(it, dict):
                    rows.append(it)
        keys = []
        seen = set()
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    keys.append(k)
                    seen.add(k)
        data = _xlsx_bytes("Resultado", keys or ["resultado"], rows or [{"resultado": text or ""}])
        headers = {"Content-Disposition": 'attachment; filename="siifa_cargue_glosas_ultimo.xlsx"'}
        return Response(
            data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    headers = {"Content-Disposition": 'attachment; filename="siifa_cargue_glosas_ultimo.json"'}
    return Response(text or "", mimetype="application/json; charset=utf-8", headers=headers)


@app.route("/cargue/glosas/plantilla.xlsx", methods=["GET"])
def cargue_glosas_plantilla():
    token = _require_token()
    if not token:
        return redirect("/login")
    headers = [
        "idFactura",
        "numeroFactura",
        "nitEmisor",
        "idSeguimientoTipoCodigoGlosa",
        "fechaFormulacion",
        "valorGlosa",
        "observacion",
    ]
    rows = [
        {
            "idFactura": None,
            "numeroFactura": "FEMF9787",
            "nitEmisor": "900243869",
            "idSeguimientoTipoCodigoGlosa": "CO2301",
            "fechaFormulacion": "2026-03-25T00:00:00Z",
            "valorGlosa": 17500,
            "observacion": "Glosa 1",
        }
    ]
    data = _xlsx_bytes("Glosas", headers, rows)
    resp_headers = {"Content-Disposition": 'attachment; filename="siifa_plantilla_glosas.xlsx"'}
    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=resp_headers,
    )


if __name__ == "__main__":
    host = os.environ.get("WEB_HOST", "0.0.0.0")
    port = int(os.environ.get("WEB_PORT", "5000"))
    debug = os.environ.get("WEB_DEBUG", "").strip() in ("1", "true", "True", "yes", "YES")
    app.run(host=host, port=port, debug=debug, use_reloader=debug)

